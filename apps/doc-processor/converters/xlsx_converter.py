from openpyxl import load_workbook

def extract_text(file_path: str) -> str:
    """Извлекает текст из XLSX файла"""
    try:
        wb = load_workbook(file_path, data_only=True)
        text = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text.append(f"=== {sheet_name} ===\n")
            
            for row in sheet.iter_rows(values_only=True):
                row_text = []
                for cell in row:
                    if cell is not None:
                        row_text.append(str(cell))
                if row_text:
                    text.append(" | ".join(row_text))
            
            text.append("\n")
        
        return "\n".join(text)
    except Exception as e:
        raise Exception(f"Failed to extract XLSX text: {str(e)}")


